import argparse
import csv
import json
import os
import re

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def parse_args():
    parser = argparse.ArgumentParser(description='log_name')
    parser.add_argument(
        'method', type=str, help='method name, used in csv/xlsx header')
    parser.add_argument(
        '--epoch',
        type=int,
        default=25,
        required=False,
        help='train_epoch, used for checking whether training completed')
    parser.add_argument(
        '--work-dirs',
        type=str,
        default='work_dirs/',
        required=False,
        help='directory for saving results')
    parser.add_argument(
        '--origin',
        type=str,
        default=False,
        required=False,
        help='excel with datasets in the order of execution ')
    args = parser.parse_args()

    return args


def write_csv(datas, args):
    num = 0
    fail_num = 0
    none_exist_num = 0
    fail = []
    none_exist = []
    latest_time = 0
    with open('scripts/labels_names.json') as f:
        label = json.load(f)
    for dataset in sorted(os.listdir(datas)):
        print(f'\ndataset={dataset}, index={num}')
        num += 1
        with open(
                os.path.join(datas, dataset, 'train/_annotations.coco.json'),
                'r') as f:
            image = json.load(f)
            num_train = len(image['images'])  # get number of train images
        with open(
                os.path.join(datas, dataset, 'valid/_annotations.coco.json'),
                'r') as f:
            image = json.load(f)
            num_valid = len(image['images'])  # get number of valid images
        for index in label:
            if index['name'] == dataset:
                category = index['category']  # get category of dataset
                class_num = len(index['classes'].keys())

        # determine whether the dataset directory exists
        try:
            dirs = [
                os.path.join(args.work_dirs, dataset, d)
                for d in os.listdir(os.path.join(args.work_dirs, dataset))
                if os.path.isdir(os.path.join(args.work_dirs, dataset, d))
            ]
            dirs.sort(key=os.path.getmtime)

            latest_dir = dirs[-1]
            latest_log_name = latest_dir.split('/')[-1]
            if int(latest_log_name) > int(latest_time):
                latest_time = latest_log_name
            print('time=' + latest_log_name)
            latest_log = latest_dir + f'/{latest_log_name}.log'
            with open(latest_log, 'r') as f:
                log = f.read()
            print(latest_log)

            complete_flag = re.findall(
                r'Epoch\(val\) \[{}\]\[\d+/\d+\]'.format(args.epoch),
                log)  # find log of args.epoch's validating process

            # Check whether the training is complete
            if not complete_flag:
                fail_num += 1
                fail.append(dataset)
                print('-------------------------------------')
                print(f'{dataset} train failed!')
                print(f'{fail_num} dataset failed!')
                print('-------------------------------------')
                key_value = [
                    dataset, category, class_num, num_train, num_valid, '', '',
                    '', '', ''
                ]
            else:
                """match result."""
                match_all = re.findall(
                    r'The best checkpoint with ([\d.]+) '
                    r'coco/bbox_mAP at ([\d.]+) epoch', log)
                if match_all:
                    match = match_all[-1]
                    best_epoch = match[-1]
                    print(f'best_epoch={best_epoch}')
                    # find best result
                    match_AP = re.findall(
                        r'\[{}\]\[\d+/\d+\]    coco/bbox_mAP: (-?\d+\.?\d*)  coco/bbox_mAP_50: (-?\d+\.?\d*)  coco/bbox_mAP_75: -?\d+\.?\d*  coco/bbox_mAP_s: (-?\d+\.?\d*)  coco/bbox_mAP_m: (-?\d+\.?\d*)  coco/bbox_mAP_l: (-?\d+\.?\d*)'  # noqa
                        .format(best_epoch),
                        log)
                    print(f'match_AP={match_AP}')

                    key_value = [
                        dataset, category, class_num, num_train, num_valid
                    ]
                    key_value.extend(match_AP[0])
                else:
                    print('----------------- --------------------------')
                    print('log has no result!')
                    print('----------------------------------------------')
                    key_value = [
                        dataset, category, class_num, num_train, num_valid, '',
                        '', '', '', ''
                    ]
        except RuntimeError:
            print(f"{dataset} directory doesn't exist!")
            none_exist_num += 1
            none_exist.append(dataset)
            key_value = [
                dataset, category, num_train, num_valid, '', '', '', '', ''
            ]

        if num == 1:  # generate headers
            result_csv = os.path.join(args.work_dirs,
                                      f'{latest_log_name}_eval.csv')
            print(result_csv)
            with open(result_csv, mode='w') as f:
                writer = csv.writer(f)
                header1 = [
                    'Dataset', 'Category', 'Classes', 'Images', 'Images',
                    args.method, args.method, args.method, args.method,
                    args.method
                ]
                writer.writerow(header1)
            with open(result_csv, mode='a') as f:
                writer = csv.writer(f)
                header2 = [
                    'Dataset', 'Category', 'Classes', 'train', 'valid', 'mAP',
                    'mAP50', 'mAP_s', 'mAP_m', 'mAP_l'
                ]
                writer.writerow(header2)
                writer.writerow(key_value)

        else:
            with open(result_csv, mode='a') as f:
                writer = csv.writer(f)
                writer.writerow(key_value)

    return result_csv, fail, fail_num, \
        none_exist, none_exist_num, os.path.join(
         args.work_dirs, latest_time[4:])


def wb_align(file, pair_ls):
    # adjust format of .xlsx file
    wb = load_workbook(file)
    ws = wb.active
    for pair in pair_ls:
        ws.merge_cells(f'{pair[0]}:{pair[1]}')
        ws[f'{pair[0]}'].alignment = Alignment(
            horizontal='center', vertical='center')
    wb.save(file)


def sort_excel(in_csv, out_xlsx):
    # read csv with two headers then convert it to xlsx,
    # sort it by category name & dataset name
    df = pd.read_csv(in_csv)
    df_sorted = df.iloc[1:].sort_values(by=['Category', 'Dataset'])
    df_sort = pd.concat([df.iloc[:1], df_sorted])
    df_sort.to_excel(out_xlsx, index=False)


def sum_excel(in_csv, out_xlsx):
    # read csv with two headers then convert it to xlsx,
    # get total number of train&valid images and mean of results
    df = pd.read_csv(in_csv)
    df.insert(2, 'dataset', pd.Series([]))
    df = df.iloc[:, 1:]
    average = df.iloc[1:].groupby('Category')  # group by category name
    df_new = df.iloc[0:1, :]
    num = 0
    for key, value in average:
        num += 1
        df_cate = [key]
        for i in range(1, 10):
            if i == 1:
                df_cate.append(len(value))
            elif i != 1 and i < 5:
                df_cate.append(value.iloc[:, i].astype(float).sum())
            else:
                # import pdb; pdb.set_trace()
                df_cate.append(
                    format(
                        value.iloc[:, i].astype(float).replace(
                            '', np.nan).replace(-1.0000, np.nan).mean(),
                        '.4f'))

        # import pdb;pdb.set_trace()
        df_new.loc[len(df_new)] = df_cate

    df_cate = ['total']  # final row = 'total'
    for i in range(1, 10):
        if i < 5:
            df_cate.append(df_new.iloc[1:, i].astype(float).sum())
        else:
            df_cate.append(
                format(
                    df_new.iloc[1:, i].astype(float).replace('',
                                                             np.nan).mean(),
                    '.4f'))
    df_new.loc[len(df_new) + 1] = df_cate
    df_new.to_excel(out_xlsx, float_format='%.4f', index=False)


def main():
    args = parse_args()

    result_csv, fail, fail_num, none_exist, \
        none_exist_num, latest_time = write_csv('rf100/', args)

    os.rename(result_csv, latest_time + '_eval.csv')
    result_csv = latest_time + '_eval.csv'

    # write excel in the order of execution
    if args.origin:
        df = pd.read_csv(result_csv)
        result_xlsx_detail = '{}_origin.xlsx'.format(latest_time)
        if os.path.exists(result_xlsx_detail):
            os.remove(result_xlsx_detail)
        print(f'\n{result_xlsx_detail} created!\n')
        df.to_excel(result_xlsx_detail)
        wb_align(result_xlsx_detail, [['E1', 'F1'], ['G1', 'K1']])

    # write excel in the order of category&dataset name
    result_xlsx_sort = '{}_detail.xlsx'.format(latest_time)
    result_xlsx_sum = '{}_sum.xlsx'.format(latest_time)
    if os.path.exists(result_xlsx_sum):
        os.remove(result_xlsx_sum)

    # sortec by category name
    sort_excel(result_csv, result_xlsx_sort)
    wb_align(result_xlsx_sort, [['D1', 'E1'], ['F1', 'J1']])

    # sum of each category
    sum_excel(result_csv, result_xlsx_sum)
    wb_align(
        result_xlsx_sum,
        [['A1', 'A2'], ['B1', 'B2'], ['C1', 'C2'], ['D1', 'E1'], ['F1', 'J1']])

    # save fail
    print(f'sum_file = {result_xlsx_sum}')
    ''' generate .txt file '''
    print(f'{none_exist_num} datasets were not trained:\n{none_exist}\n')
    print(f'{fail_num} training failed:\n{fail}\n')

    fail_txt = os.path.join(args.work_dirs, 'failed_dataset_list.txt')
    with open(fail_txt, 'w') as f:
        pass
    with open(fail_txt, 'a') as f:
        for item in none_exist:
            f.write(f'{item}\n')
        for item in fail:
            f.write(f'{item}\n')

    print(f'all {fail_num + none_exist_num} untrained datasets '
          f'have been logged in {fail_txt}!')


if __name__ == '__main__':
    main()
